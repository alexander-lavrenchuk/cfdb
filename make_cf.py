# from config.db_connection_config import host, port, user, password, db_name
import pandas as pd
from py.mysql_db import select_data, get_min_max_dates
from py.card51 import get_periods
# from config.config import abs_accurancy, reports_directory_path
from config.config import reports_directory_path
from os.path import join, exists
from os import remove
import openpyxl
from openpyxl.utils import get_column_letter
# from collections.abc import Callable
from collections import namedtuple
from time import time

field_names = [
    'left_field',
    'right_table',
    'id',
    'right_field',
    'width',
]

ShowField = namedtuple(
    typename='ShowFields',
    field_names=field_names,
    defaults=(None,) * len(field_names)
)

ReportParams = namedtuple(
    typename='ReportParam',
    field_names=[
        'file_name',
        'group_by',
        'show_fields'
    ]
)


def move_column(data_frame: pd.DataFrame,
    column_name: str, loc: int):

    col = data_frame.pop(column_name)
    data_frame.insert(
        loc=loc,
        column=column_name,
        value=col)


def to_quoted_fields_string(fields: list[str]) -> str:
    quoted_fields = \
        ['`' + field + '`' for field in fields]
    quoted_fields_string = ', '.join(quoted_fields)
    return quoted_fields_string


def select_sum_group_by_fields(
        table_name: str,
        fields: list[str],
        sum_field_name: str) -> pd.DataFrame:
    group_by_fields = to_quoted_fields_string(fields)
    sql_query = f"""
        SELECT {group_by_fields},
            SUM(`amount`) AS {sum_field_name}
        FROM {table_name}
        GROUP BY {group_by_fields};
    """
    df = select_data(sql_query)
    df[fields] = df[fields].fillna(-1)
    return df


def cf_to_excel(
        file_name: str,
        group_by: list[str],
        show_fields: list[ShowField],
        value_cell_width: int,
        perform_formatting: bool=True) -> None:

    # appending_mode = False
    activity_types = ('op', 'inv', 'fin', 'equity', 'vgo')
    transaction_directions = ('ins', 'outs')

    file_path = join(reports_directory_path, file_name)
    if exists(file_path):
        remove(file_path)

    for activity_type in activity_types:
        for transaction_direction in transaction_directions:
            table_name = f"{transaction_direction}_{activity_type}"
            
            fields=['period_name'] + group_by
            sum_field_name = 'sum_by_all'

            df = select_sum_group_by_fields(
                table_name=table_name,
                fields=fields,
                sum_field_name=sum_field_name
            )

            df = pd.pivot_table(
                data=df,
                values=sum_field_name,
                index=group_by,
                columns='period_name',
                aggfunc='sum',
                fill_value=0,
                margins=False,
                sort=False)

            df = df.reset_index()

            dfs = []
            columns_to_sort = []
            for field_nums in range(len(group_by)):
                fields = group_by[:field_nums + 1]
                sum_field_name = 'sum_by_' + str(field_nums + 1)
                columns_to_sort.append(sum_field_name)

                dft = select_sum_group_by_fields(
                    table_name=table_name,
                    fields=fields,
                    sum_field_name=sum_field_name
                )

                df = df.merge(
                    right=dft[[*fields, sum_field_name]],
                    how='left',
                    left_on=fields,
                    right_on=fields
                )

                dfs.append(dft)

            df.sort_values(
                by=columns_to_sort,
                ascending=[False] * len(group_by),
                inplace=True)

            columns_to_drop = group_by.copy()

            for field in show_fields:
                if not field.right_table:
                    columns_to_drop.remove(field.left_field)
                    continue

                if type(field.right_table) == str:
                    right_table = field.right_table
                else:
                    right_table = field.right_table[transaction_direction == 'outs']

                sql_query = f"""
                    SELECT
                        `{field.id}`,
                        `{field.right_field}`
                    FROM `{right_table}`;
                """
                dft = select_data(sql_query)
                df = df.merge(
                    right=dft,
                    how='left',
                    left_on=field.left_field,
                    right_on=field.id
                )
                df.drop(columns=field.id, inplace=True)
                df[field.right_field] = df[field.right_field].fillna('')

            df.drop(columns=columns_to_drop, inplace=True)
            df.drop(columns=columns_to_sort, inplace=True)

            df_periods = select_data('SELECT period_name FROM periods;')
            all_periods_set = set(df_periods.period_name)
            missing_periods_set = all_periods_set - set(df.columns)
            if len(missing_periods_set) > 0:
                for period in missing_periods_set:
                    df.insert(loc=df.shape[1], column=period, value=0)

            # columns_num_to_sum = df.shape[1]-len(show_fields)

            df.sort_index(
                axis=1,
                inplace=True)
            
            field_names = []
            for field in show_fields:
                field_names.append((field.right_field, field.left_field)[field.right_field is None])
            df.set_index(keys=field_names, inplace=True)

            df_total_row = df.sum(axis=0).to_frame().T
            for field_name in field_names:
                df_total_row[field_name] = pd.Series('All')

            df_total_row.set_index(keys=field_names, inplace=True)

            df = pd.concat([df_total_row, df])

            df['All'] = df.sum(axis=1)
            
            df.reset_index(names=field_names, inplace=True)

            writer_mode = 'a' if exists(file_path) else 'w'
            if transaction_direction == 'ins':
                sheet_behaviour = None
                header = True
                startrow = 1
            else:
                sheet_behaviour = 'overlay'
                header = False
                startrow = writer.sheets[activity_type].max_row + 1

            with pd.ExcelWriter(
                file_path,
                mode=writer_mode,
                if_sheet_exists=sheet_behaviour) as writer:  

                df.to_excel(
                    writer,
                    sheet_name=activity_type,
                    float_format='%.2f',
                    index=False,
                    header=header,
                    startrow=startrow,
                    freeze_panes=(1, 4)
                )
            
        if perform_formatting:
            format_sheet(
                wb_path=file_path,
                show_fields=show_fields,
                value_cell_width=value_cell_width
            )

    print(f'CF model saved to {file_path}')


def format_sheet(
        wb_path: str,
        show_fields: list[ShowField],
        value_cell_width: int
        ) -> None:

    wb = openpyxl.load_workbook(wb_path)

    for sh_name in wb.sheetnames:
        sh = wb[sh_name]

        min_row = 1
        max_row = sh.max_row
        min_col = len(show_fields) + 1
        max_col = sh.max_column
        cell_range = sh.iter_cols(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col
        )

        for col in range(1, min_col):
            width = show_fields[col - 1].width
            if width > 0:
                sh.column_dimensions[get_column_letter(col)].width = width
            else:
                sh.column_dimensions[get_column_letter(col)].hidden = True

        for row in range(min_row, max_row + 1):
            for col_idx in range(1, min_col):
                sh[chr(ord('A') + col_idx - 1) + str(row)].alignment = \
                    openpyxl.styles.Alignment(horizontal='left')

        for col in range(min_col, max_col + 1):
            sh.column_dimensions[get_column_letter(col)].width = value_cell_width

        for row in cell_range:
            for cell in row:
                cell.number_format = '#,##0.00'

    wb.save(wb_path)
    wb.close()


def fix_time(start_time: float) -> str:
    duration_in_seconds = time() - start_time
    hh, ss = divmod(duration_in_seconds, 3600)
    mm, ss = divmod(ss, 60)
    ss, ms = divmod(ss, 1)
    hh, mm, ss, ms = map(int, [hh, mm, ss, 1000 * ms])
    time_elapsed_string = f"{(hh):02d}:{(mm):02d}:{(ss):02d}.{(ms):03d}"
    return time_elapsed_string


if __name__ == '__main__':
    start_time = time()
    
    account_id = ShowField(
        left_field='account_id',
        width=8
    )
    
    account_name = ShowField(
        left_field='account_id',
        right_table='accounts',
        id='id',
        right_field='account_name',
        width=32
    )
    
    entity_1C_name = ShowField(
        left_field='entity_id',
        right_table='entities',
        id='id',
        right_field='entity_1C_name',
        width=32
    )
    
    inn = ShowField(
        left_field='entity_id',
        right_table='entities',
        id='id',
        right_field='inn',
        width=14
    )
    
    full_name = ShowField(
        left_field='entity_id',
        right_table='entities',
        id='id',
        right_field='full_name',
        width=32
    )
    
    article_name = ShowField(
        left_field='article_id',
        right_table=['income_articles', 'outcome_articles'],
        id='id',
        right_field='article_name',
        width=32
    )

    value_cell_width = 18

    params = []

    report_params = ReportParams(
        file_name = 'cf_by_entity.xlsx',
        group_by = ['entity_id'],
        show_fields = [entity_1C_name, inn]
    )
    params.append(report_params)

    report_params = ReportParams(
        file_name = 'cf_by_account.xlsx',
        group_by = ['account_id'],
        show_fields = [account_id, account_name]
    )
    params.append(report_params)

    report_params = ReportParams(
        file_name = 'cf_by_article.xlsx',
        group_by = ['article_id'],
        show_fields = [article_name]
    )
    params.append(report_params)

    report_params = ReportParams(
        file_name = 'cf_by_account_entity.xlsx',
        group_by = ['account_id', 'entity_id'],
        show_fields = [account_id, account_name, entity_1C_name, inn]
    )
    params.append(report_params)

    report_params = ReportParams(
        file_name = 'cf_by_entity_account.xlsx',
        group_by = ['entity_id', 'account_id'],
        show_fields = [entity_1C_name, inn, account_id, account_name]
    )
    params.append(report_params)

    report_params = ReportParams(
        file_name = 'cf_by_article_entity.xlsx',
        group_by = ['article_id', 'entity_id'],
        show_fields = [article_name, entity_1C_name, inn]
    )
    params.append(report_params)

    report_params = ReportParams(
        file_name = 'cf_by_entity_article.xlsx',
        group_by = ['entity_id', 'article_id'],
        show_fields = [entity_1C_name, inn, article_name]
    )
    params.append(report_params)

    report_params = ReportParams(
        file_name = 'cf_by_account_article.xlsx',
        group_by = ['account_id', 'article_id'],
        show_fields = [account_id, account_name, article_name]
    )
    params.append(report_params)

    report_params = ReportParams(
        file_name = 'cf_by_article_account.xlsx',
        group_by = ['article_id', 'account_id'],
        show_fields = [article_name, account_id, account_name]
    )
    params.append(report_params)

    for report_params in params:
        print(f"Generating report {report_params.file_name}")
        report_time = time()
        cf_to_excel(
            file_name=report_params.file_name,
            group_by=report_params.group_by,
            show_fields=report_params.show_fields,
            value_cell_width=value_cell_width,
            perform_formatting=True
        )
        time_elapsed = fix_time(report_time)
        print(f"Done in {time_elapsed} seconds")
        print()

    print(f"Total time elapsed {fix_time(start_time)} seconds")

